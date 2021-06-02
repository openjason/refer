#python 操作openpyxl导出Excel 设置单元格格式以及合并处理
#贴上一个例子，里面设计很多用法，根据将相同日期的某些行合并处理。
from openpyxl import Workbook                                                                                                                                 
from openpyxl.styles import Font, Fill, Alignment, Border, Side, PatternFill                                                                                  
from handlers.boss_accountant import PbOrderManageBase                                                                                                        
from handlers.base.pub_func import ConfigFunc                                                                                                                 
from dal.models import Shop                                                                                                                                   
from dal.db_configs import DBSession                                                                                                                          
                                                                                                                                                              
                                                                                                                                                              
                                                                                                                                                              
                                                                                                                                                              
def export_excel(filename, sheetname, content_body):                                                                                                          
    """                                                                                                                                                       
    Excel表格导出                                                                                                                                                 
    :param filename: 表格名称                                                                                                                                     
    :param sheetname: 工作表名称                                                                                                                                   
    :param content_body: 内容体                                                                                                                                  
    :return: None                                                                                                                                             
    """                                                                                                                                                       
    workbook = Workbook()                                                                                                                                     
                                                                                                                                                              
    if not filename:                                                                                                                                          
        filename = "导出表格.xlsx"                                                                                                                                
                                                                                                                                                              
    workbook_sheet = workbook.active                                                                                                                          
    if not sheetname:                                                                                                                                         
        sheetname = "工作表"                                                                                                                                     
    workbook_sheet.title = sheetname                                                                                                                          
                                                                                                                                                              
    merge_dict, sheet_row_len, sheet_column_len = merge_content(content_body)                                                                                 
    print(merge_dict)                                                                                                                                         
    # 数据写入                                                                                                                                                    
    for row in content_body:                                                                                                                                  
        workbook_sheet.append(row)                                                                                                                            
                                                                                                                                                              
    # 合并处理                                                                                                                                                    
    for key in merge_dict.keys():                                                                                                                             
        merge_data = merge_dict.get(key)                                                                                                                      
        if key == "title":                                                                                                                                    
            workbook_sheet.merge_cells(start_row=merge_data[0], start_column=merge_data[1],                                                                   
                                       end_row=merge_data[2], end_column=merge_data[3])                                                                       
            workbook_sheet.merge_cells(start_row=2, start_column=merge_data[1],                                                                               
                                       end_row=2, end_column=merge_data[3])                                                                                   
            workbook_sheet['A1'].font = Font(size=20, bold=True)                                                                                              
            workbook_sheet['A1'].alignment = Alignment(horizontal='center', vertical='center')                                                                
        else:                                                                                                                                                 
            # 使用sum求值                                                                                                                                         
            workbook_sheet.cell(row=merge_data[0] + 3, column=12).value = '=SUM({}:{})'.format(                                                               
                format_value(str(merge_data[0] + 3), 10), format_value(str(merge_data[1] + 3), 10))                                                           
            workbook_sheet.cell(row=merge_data[0] + 3, column=14).value = '=SUM({}:{})'.format(                                                               
                format_value(str(merge_data[0] + 3), 11), format_value(str(merge_data[1] + 3), 11))                                                           
            workbook_sheet.cell(row=merge_data[0] + 3, column=13).value = '=({}-{})'.format(                                                                  
                format_value(str(merge_data[0] + 3), 12), format_value(str(merge_data[0] + 3), 14))                                                           
                                                                                                                                                              
            for i in [2,12, 13, 14]:                                                                                                                          
                workbook_sheet.merge_cells(start_row=merge_data[0]+3, start_column=i,                                                                         
                                           end_row=merge_data[1]+3, end_column=i)                                                                             
    # 合计求和                                                                                                                                                    
    for i in [12, 13, 14]:                                                                                                                                    
        workbook_sheet.cell(row=sheet_row_len, column=i).value = '=SUM({}:{})'.format(                                                                        
            format_value(3, i), format_value(sheet_row_len - 1, i))                                                                                           
                                                                                                                                                              
    # 单元格底色                                                                                                                                                   
    last_row = workbook_sheet[sheet_row_len]                                                                                                                  
    for each_cell in last_row:                                                                                                                                
        each_cell.fill = PatternFill("solid", fgColor="00CDCD")                                                                                               
                                                                                                                                                              
    # 边框设置                                                                                                                                                    
    for each_common_row in workbook_sheet.iter_rows("A1:{}".format(format_value(sheet_row_len, sheet_column_len))):                                           
        for each_cell in each_common_row:                                                                                                                     
            each_cell.border = Border(left=Side(style='thin', color='000000'),                                                                                
                                      right=Side(style='thin', color='000000'),                                                                               
                                      top=Side(style='thin', color='000000'),                                                                                 
                                      bottom=Side(style='thin', color='000000')                                                                               
                                      )                                                                                                                       
    workbook_sheet.column_dimensions['B'].width = 15                                                                                                          
    workbook_sheet.column_dimensions['C'].width = 20                                                                                                          
    workbook.save(filename)                                                                                                                                   
                                                                                                                                                              
                                                                                                                                                              
def merge_content(content_body):                                                                                                                              
    """                                                                                                                                                       
    合并统计                                                                                                                                                      
    :param content_body: 数据体                                                                                                                                  
    :return: 合并字典                                                                                                                                             
    """                                                                                                                                                       
    sheet_column_len = len(content_body[3])                                                                                                                   
    sheet_row_len = len(content_body)                                                                                                                         
    merge_dict = {}                                                                                                                                           
                                                                                                                                                              
    data_content = content_body[3:-1]                                                                                                                         
                                                                                                                                                              
    merge_dict["title"] = (1, 1, 1, sheet_column_len)                                                                                                         
                                                                                                                                                              
    current_data = data_content[0][1]                                                                                                                         
    current_row = 0                                                                                                                                           
    start_row = 1                                                                                                                                             
    end_row = 0                                                                                                                                               
                                                                                                                                                              
    for data in data_content:                                                                                                                                 
        current_row += 1                                                                                                                                      
        x = data[1]                                                                                                                                           
        if data[1] == current_data:                                                                                                                           
                                                                                                                                                              
            merge_dict[data[1]] = (start_row, current_row)                                                                                                    
        else:                                                                                                                                                 
                                                                                                                                                              
            merge_dict[data[1]] = (current_row, current_row)                                                                                                  
            current_data = data[1]                                                                                                                            
                                                                                                                                                              
            start_row = current_row                                                                                                                           
                                                                                                                                                              
    return merge_dict, sheet_row_len, sheet_column_len                                                                                                        
                                                                                                                                                              
                                                                                                                                                              
def format_value(row, column):                                                                                                                                
    """数字转ABC                                                                                                                                                 
    """                                                                                                                                                       
    change_dict = {                                                                                                                                           
        1: "A", 2: "B", 3: "C", 4: "D", 5: "E", 6: "F", 7: "G", 8: "H", 9: "I", 10: "J",                                                                      
        11: "K", 12: "L", 13: "M", 14: "N", 15: "O", 16: "P", 17: "Q", 18: "R", 19: "S", 20: "T",                                                             
        21: "U", 22: "V", 23: "W", 24: "X", 25: "Y", 26: "Z",                                                                                                 
    }                                                                                                                                                         
    column = change_dict.get(column)                                                                                                                          
    return str(column)+str(row)                                                                                                                               
                                                                                                                                                              
                                                                                                                                                              
def export_func_new(args, session, shop_id):                                                                                                                  
    # check_time = 0                                                                                                                                          
    # debtor_id = 2884                                                                                                                                        
    # debtor_name: 肖小菜                                                                                                                                        
    # end_date:                                                                                                                                               
    # start_date: 2019 - 07                                                                                                                                   
    # statistic_date: 3                                                                                                                                       
    # data_type: 1                                                                                                                                            
    data_content = []                                                                                                                                         
    check_time = 0                                                                                                                                            
    from_date = "2019-07"                                                                                                                                     
    to_date = ""                                                                                                                                              
    debtor_name = "肖小菜"                                                                                                                                       
                                                                                                                                                              
    if_success, query_data, *_ = PbOrderManageBase.common_get_credit_stream(args, session, shop_id, export=True,                                              
                                                                            need_sum=False, check_time=check_time                                             
    )                                                                                                                                                         
    if not if_success:                                                                                                                                        
        raise ValueError(query_data)                                                                                                                          
    fee_text = ConfigFunc.get_fee_text(session, shop_id)                                                                                                      
    get_weight_unit_text = ConfigFunc.get_weight_unit_text(session, shop_id)                                                                                  
                                                                                                                                                              
    # 表店铺、客户名称                                                                                                                                                
    shop_name = session.query(Shop.shop_name).filter_by(id=shop_id).first()                                                                                   
    data_content.append([shop_name[0]])                                                                                                                       
    data_content.append(["客户：{}".format(debtor_name)])                                                                                                        
    # 表头                                                                                                                                                      
    fee_text_total = '{}小计'.format(fee_text)                                                                                                                  
    header_content = [                                                                                                                                        
        "序号", "日期", "货品名", "数量", "重量/{}".format(get_weight_unit_text), "单价", "货品小记", "押金小计", fee_text_total,                                                  
        "赊账金额","待还款", "赊账小记", "已还款", "待还款小计"                                                                                                                  
    ]                                                                                                                                                         
    file_name_begin = "客户还款"                                                                                                                                  
    data_content.append(header_content)                                                                                                                       
    # 还款数据                                                                                                                                                    
    index_num = 0                                                                                                                                             
    for single_data in query_data:                                                                                                                            
        index_num += 1                                                                                                                                        
        sales_time = single_data.get("sales_time", "")                                                                                                        
        if sales_time:                                                                                                                                        
            sales_time = sales_time.split(" ")[0]                                                                                                             
                                                                                                                                                              
        _payback_money = single_data["unpayback_money"]                                                                                                       
        single_content = [index_num,                                                                                                                          
                          sales_time,                                                                                                                         
                          single_data["only_goods_name"],                                                                                                     
                          single_data["commission_mul"],                                                                                                      
                          single_data["sales_num"],                                                                                                           
                          "%s元/%s" % (single_data["fact_price"],                                                                                              
                                      single_data["goods_unit"]),                                                                                             
                          single_data["goods_total"],                                                                                                         
                          single_data["commission_mul"],                                                                                                      
                          single_data["deposit_total"],                                                                                                       
                          single_data["credit_cent"],                                                                                                         
                          _payback_money,                                                                                                                     
                          0,                                                                                                                                  
                          0,                                                                                                                                  
                          0]                                                                                                                                  
        data_content.append(single_content)                                                                                                                   
    # 表尾合计                                                                                                                                                    
    data_content.append(["合计"])                                                                                                                               
    config = ConfigFunc.get_config(session, shop_id)                                                                                                          
    if not config.enable_deposit:                                                                                                                             
        index_deposit_total = data_content[0].index("押金小计")                                                                                                   
        for data in data_content:                                                                                                                             
            data.pop(index_deposit_total)                                                                                                                     
                                                                                                                                                              
    if not config.enable_commission:                                                                                                                          
        index_commission_total = data_content[0].index(fee_text_total)                                                                                        
        for data in data_content:                                                                                                                             
            data.pop(index_commission_total)                                                                                                                  
                                                                                                                                                              
    file_name = "{}流水记录导出_{}~{}.xlsx".format(file_name_begin, from_date, to_date)                                                                             
    return file_name, data_content                                                                                                                            
                                                                                                                                                              
                                                                                                                                                              
if __name__ == "__main__":                                                                                                                                    
    filename = "测试打印表格.xlsx"                                                                                                                                  
    sheetname = "工作表2"                                                                                                                                        
                                                                                                                                                              
                                                                                                                                                              
                                                                                                                                                              
    session = DBSession()                                                                                                                                     
    args = {                                                                                                                                                  
        "check_time": 0,                                                                                                                                      
        "debtor_id": 2884,                                                                                                                                    
        "debtor_name": "肖小菜",                                                                                                                                 
        "start_date": "2019-07",                                                                                                                              
        "statistic_date": 3,                                                                                                                                  
        "data_type": 1                                                                                                                                        
    }                                                                                                                                                         
    filename, content_body = export_func_new(args, session, 104)                                                                                              
                                                                                                                                                              
                                                                                                                                                              
    # filename = "测试打印表格.xlsx"                                                                                                                                
    # sheetname = "工作表2"                                                                                                                                      
    # content_body = []                                                                                                                                       
    # content_body.append(["打印表格表头"])                                                                                                                         
    # content_body.append(["客户：肖某某"])                                                                                                                         
    # content_body.append(["日期", "货品销售", "自营销售", "代卖销售", "联营销售", "总价"])                                                                                       
    # content_body.append(["1", "2", "3.1", "4.1", "5.1", "5.1"])                                                                                             
    # content_body.append(["1", "2", "3.1", "4.1", "5.1", "5.1"])                                                                                             
    # content_body.append(["1", "2", "3.1", "4.1", "5.1", "5.1"])
