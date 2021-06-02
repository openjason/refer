import os							#找文件目录
import win32com.client as win32 	#操作excel文件
from tqdm import tqdm 				#进度条显示
from openpyxl import load_workbook # 读取时导入这个
from openpyxl.styles import Font, Alignment #设置单元格格式
from openpyxl.utils import get_column_letter, column_index_from_string
from openpyxl.styles import PatternFill, Border, Side, Alignment, Protection, Font
from copy import copy
path=input('输入整理前原始路径： ')
if path=="":
	path=os.getcwd()
xlsx_lists=[]
xls_lists=[]
for file in os.listdir(path):
 filename=os.path.join(path,file)
 if os.path.isfile(filename):    #是目录
  if filename.endswith(".xls"):
  	xls_lists.append(filename)
  if filename.endswith(".xlsx"):
  	xlsx_lists.append(filename)
source_file='原始数据.xlsx'
if os.path.exists(os.path.join(os.getcwd(),source_file)):
	os.remove(os.path.join(os.getcwd(),source_file))
choose="1"
excel = win32.gencache.EnsureDispatch('Excel.Application')
# while choose not in "1|2":
#  choose =input("xls转为xlsx：1 xlsx转为xls：2 ")
if choose=="1":
	with tqdm(total=len(xls_lists),desc='写文件数 ',leave=True,unit='个',unit_scale=True,mininterval=0.5,bar_format=None) as pbar:
		for xls_list in xls_lists:
			pbar.update(1)
			wb = excel.Workbooks.Open(xls_list)
			wb.SaveAs(xls_list+"x", FileFormat = 51) #FileFormat = 51 is for .xlsx extension
			wb.Close()        #FileFormat = 56 is for .xls extension
		pbar.close()
else:
	with tqdm(total=len(xls_lists),desc='写文件数 ',leave=True,unit='个',unit_scale=True,mininterval=0.5,bar_format=None) as pbar:
		for xlsx_list in xlsx_lists:
			pbar.update(1)
			wb = excel.Workbooks.Open(xlsx_list)
			wb.SaveAs(xlsx_list[0:len(xlsx_list)-1], FileFormat = 56) #FileFormat = 51 is for .xlsx extension
			wb.Close() 
		pbar.close()
excel.Application.Quit()
tag_file='拆分后表.xlsx'
totaldata=pd.DataFrame()
writer=pd.ExcelWriter(tag_file)
totaldata.to_excel(writer, 'sheet')
writer.save()
book = load_workbook(tag_file)   #能写入已存在表中
wb = load_workbook('原始数据.xlsx')
for sheet in wb.sheetnames:
	print(sheet)
	wbsheet=wb[sheet]
	for num in range(3):
		name=wbsheet.cell(1,num*15+10).value
		wbsheet_new = book.create_sheet(name,0)
		wm=list(wbsheet.merged_cells) #开始处理合并单元格形式为“(<CellRange A1：A4>,)，替换掉(<CellRange 和 >,)' 找到合并单元格
		#print (list(wm))
		if len(wm)>0 :
			for i in range(0,len(wm)):
				cell2=str(wm[i]).replace('(<CellRange ','').replace('>,)','')
				#print("MergeCell : %s" % cell2)
				wbsheet_new.merge_cells(cell2)
		for rows in range(40):
			wbsheet_new.row_dimensions[rows+1].height = wbsheet.row_dimensions[rows+1].height 
			for col in range(14):
				wbsheet_new.column_dimensions[get_column_letter(col+1)].width = wbsheet.column_dimensions[get_column_letter(col+1)].width
				wbsheet_new.cell(row=rows+1,column=col+1,value=wbsheet.cell(rows+1,num*15+col+1).value)
				if wbsheet.cell(rows+1,num*15+col+1).has_style:	#拷贝格式
					wbsheet_new.cell(row=rows+1,column=col+1).font = copy(wbsheet.cell(rows+1,num*15+col+1).font)
					wbsheet_new.cell(row=rows+1,column=col+1).border = copy(wbsheet.cell(rows+1,num*15+col+1).border)
					wbsheet_new.cell(row=rows+1,column=col+1).fill = copy(wbsheet.cell(rows+1,num*15+col+1).fill)
					wbsheet_new.cell(row=rows+1,column=col+1).number_format = copy(wbsheet.cell(rows+1,num*15+col+1).number_format)
					wbsheet_new.cell(row=rows+1,column=col+1).protection = copy(wbsheet.cell(rows+1,num*15+col+1).protection)
					wbsheet_new.cell(row=rows+1,column=col+1).alignment = copy(wbsheet.cell(rows+1,num*15+col+1).alignment)
wb.close()
book.save('拆分后表.xlsx')
book.close()
