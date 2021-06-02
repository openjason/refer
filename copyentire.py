#!/usr/bin/env python
# coding=utf-8
import openpyxl
from copy import copy

def copy_same_file_path_sheet_with_format(targetsheet, sourcesheet, beginrow=0,begincol=0,endrow=999,endcol=99,offsetrow=0,offsetcol=0): 
    #targetbk:一个workbook，将会新建一个sheet，openpyxl.load_workbook(excelFile2)
    #sourcesheet：一个已存在的wb = openpyxl.load_workbook(excelFile)；sh = wb[sheet[0]]

    #offsetrow = 6
    #offsetcol = 6
    ws = sourcesheet  
    ws2 = targetsheet
    if endrow > ws.max_row:
        max_row=ws.max_row    #最大行数 
    else:
        max_row=endrow
    if endcol > ws.max_column:
        max_column=ws.max_column    #最大列数
    else:
        max_column= endcol

    m_list=ws.merged_cells     #合并单元格的位置信息，可迭代对象（单个是一个'openpyxl.worksheet.cell_range.CellRange'对象），print后就是excel坐标信息
    mergedcellqty=len(ws.merged_cells.ranges)
    if mergedcellqty>0 :      
        cr = []
        for m_area in m_list:
            # 合并单元格的起始行坐标、终止行坐标。。。。，
            r1, r2, c1, c2 = m_area.min_row, m_area.max_row, m_area.min_col, m_area.max_col
            # 纵向合并单元格的位置信息提取出
            if r2 - r1 >= 0:
                cr.append((r1, r2, c1, c2))
                #print('符合条件%s' % str(m_area))
        for r in cr:
            ws2.merge_cells(start_row=r[0]+offsetrow, end_row=r[1]+offsetrow, start_column=r[2] + offsetcol, end_column=r[3] + offsetcol)

    for m in range(beginrow,max_row + 1):      
        ws2.row_dimensions[m].height = ws.row_dimensions[m].height      
        for n in range(begincol,1 + max_column):           
            if n<27 :               
                c=chr(n+64).upper() #ASCII字符,chr(65)='A'
            else:               
                if n < 677 :                    
                    c=chr(divmod(n,26)[0]+64)+chr(divmod(n,26)[1]+64)               
                else:                   
                    c=chr(divmod(n,676)[0]+64) + chr(divmod(divmod(n,676)[1],26)[0]+64) + chr(divmod(divmod(n,676)[1],26)[1]+64)            
            ncol = n + offsetcol
            if ncol<27 :               
                c2=chr(ncol+64).upper() #ASCII字符,chr(65)='A'
            else:               
                if ncol < 677 :                    
                    c2=chr(divmod(ncol,26)[0]+64)+chr(divmod(ncol,26)[1]+64)               
                else:                   
                    c2=chr(divmod(ncol,676)[0]+64) + chr(divmod(divmod(ncol,676)[1],26)[0]+64) + chr(divmod(divmod(ncol,676)[1],26)[1]+64)            

            i='%s%d'%(c,m)  #单元格编号          
            j='%s%d'%(c2,m + offsetrow)  #单元格编号          
            if m == 1 :                 
                pass
                ws2.column_dimensions[c].width = ws.column_dimensions[c].width         
            try:                
                cell1=ws[i]    #获取data单元格数据             
                ws2[j].value=cell1.value    #赋值到ws2单元格              
                if cell1.has_style: #拷贝格式                   
                    ws2[j].font = copy(cell1.font)                  
                    ws2[j].border = copy(cell1.border)                  
                    ws2[j].fill = copy(cell1.fill)                  
                    ws2[j].number_format = copy(cell1.number_format)                    
                    ws2[j].protection = copy(cell1.protection)                  
                    ws2[j].alignment = copy(cell1.alignment)            
            except AttributeError as e:             
                print("cell(%s) is %s" % (i,e))             
                continue    



def copy_path_sheet_with_format(targetbk, sourcesheet, beginrow=0,begincol=0,endrow=9999,endcol=9999,offsetrow=0,offsetcol=0): 
    #targetbk:一个workbook，将会新建一个sheet，openpyxl.load_workbook(excelFile2)
    #sourcesheet：一个已存在的wb = openpyxl.load_workbook(excelFile)；sh = wb[sheet[0]]

    #offsetrow = 6
    #offsetcol = 6
    ws = sourcesheet  
    ws2 = targetbk.create_sheet() 
    if endrow > ws.max_row:
        max_row=ws.max_row    #最大行数 
    else:
        max_row=endrow
    if endcol > ws.max_column:
        max_column=ws.max_column    #最大列数
    else:
        max_column= endcol

    m_list=ws.merged_cells     #合并单元格的位置信息，可迭代对象（单个是一个'openpyxl.worksheet.cell_range.CellRange'对象），print后就是excel坐标信息
    mergedcellqty=len(ws.merged_cells.ranges)
    if mergedcellqty>0 :      
        cr = []
        for m_area in m_list:
            # 合并单元格的起始行坐标、终止行坐标。。。。，
            r1, r2, c1, c2 = m_area.min_row, m_area.max_row, m_area.min_col, m_area.max_col
            # 纵向合并单元格的位置信息提取出
            if r2 - r1 >= 0:
                cr.append((r1, r2, c1, c2))
                #print('符合条件%s' % str(m_area))
        for r in cr:
            ws2.merge_cells(start_row=r[0]+offsetrow, end_row=r[1]+offsetrow, start_column=r[2] + offsetcol, end_column=r[3] + offsetcol)

    for m in range(1,max_row + 1):      
        ws2.row_dimensions[m].height = ws.row_dimensions[m].height      
        for n in range(1,1 + max_column):           
            if n<27 :               
                c=chr(n+64).upper() #ASCII字符,chr(65)='A'
            else:               
                if n < 677 :                    
                    c=chr(divmod(n,26)[0]+64)+chr(divmod(n,26)[1]+64)               
                else:                   
                    c=chr(divmod(n,676)[0]+64) + chr(divmod(divmod(n,676)[1],26)[0]+64) + chr(divmod(divmod(n,676)[1],26)[1]+64)            
            ncol = n + offsetcol
            if ncol<27 :               
                c2=chr(ncol+64).upper() #ASCII字符,chr(65)='A'
            else:               
                if ncol < 677 :                    
                    c2=chr(divmod(ncol,26)[0]+64)+chr(divmod(ncol,26)[1]+64)               
                else:                   
                    c2=chr(divmod(ncol,676)[0]+64) + chr(divmod(divmod(ncol,676)[1],26)[0]+64) + chr(divmod(divmod(ncol,676)[1],26)[1]+64)            

            i='%s%d'%(c,m)  #单元格编号          
            j='%s%d'%(c2,m + offsetrow)  #单元格编号          
            if m == 1 :                 
                ws2.column_dimensions[c].width = ws.column_dimensions[c].width         
            try:                
                cell1=ws[i]    #获取data单元格数据             
                ws2[j].value=cell1.value    #赋值到ws2单元格              
                if cell1.has_style: #拷贝格式                   
                    ws2[j].font = copy(cell1.font)                  
                    ws2[j].border = copy(cell1.border)                  
                    ws2[j].fill = copy(cell1.fill)                  
                    ws2[j].number_format = copy(cell1.number_format)                    
                    ws2[j].protection = copy(cell1.protection)                  
                    ws2[j].alignment = copy(cell1.alignment)            
            except AttributeError as e:             
                print("cell(%s) is %s" % (i,e))             
                continue    


def copy_entire_sheet_with_format(targetbk, sourcesheet): 
    #targetbk:一个workbook，将会新建一个sheet，openpyxl.load_workbook(excelFile2)
    #sourcesheet：一个已存在的wb = openpyxl.load_workbook(excelFile)；sh = wb[sheet[0]]

    ws = sourcesheet  
    ws2 = targetbk.create_sheet() 
    max_row=ws.max_row    #最大行数 
    max_column=ws.max_column    #最大列数

    m_list=ws.merged_cells     #合并单元格的位置信息，可迭代对象（单个是一个'openpyxl.worksheet.cell_range.CellRange'对象），print后就是excel坐标信息
    mergedcellqty=len(ws.merged_cells.ranges)
    if mergedcellqty>0 :      
        cr = []
        for m_area in m_list:
            # 合并单元格的起始行坐标、终止行坐标。。。。，
            r1, r2, c1, c2 = m_area.min_row, m_area.max_row, m_area.min_col, m_area.max_col
            # 纵向合并单元格的位置信息提取出
            if r2 - r1 >= 0:
                cr.append((r1, r2, c1, c2))
                #print('符合条件%s' % str(m_area))
        for r in cr:
            ws2.merge_cells(start_row=r[0], end_row=r[1], start_column=r[2], end_column=r[3])

    for m in range(1,max_row + 1):      
        ws2.row_dimensions[m].height = ws.row_dimensions[m].height      
        for n in range(1,1 + max_column):           
            if n<27 :               
                c=chr(n+64).upper() #ASCII字符,chr(65)='A'            
            else:               
                if n < 677 :                    
                    c=chr(divmod(n,26)[0]+64)+chr(divmod(n,26)[1]+64)               
                else:                   
                    c=chr(divmod(n,676)[0]+64) + chr(divmod(divmod(n,676)[1],26)[0]+64) + chr(divmod(divmod(n,676)[1],26)[1]+64)            
            i='%s%d'%(c,m)  #单元格编号          
            if m == 1 :                 
                ws2.column_dimensions[c].width = ws.column_dimensions[c].width         
            try:                
                cell1=ws[i]    #获取data单元格数据             
                ws2[i].value=cell1.value    #赋值到ws2单元格              
                if cell1.has_style: #拷贝格式                   
                    ws2[i].font = copy(cell1.font)                  
                    ws2[i].border = copy(cell1.border)                  
                    ws2[i].fill = copy(cell1.fill)                  
                    ws2[i].number_format = copy(cell1.number_format)                    
                    ws2[i].protection = copy(cell1.protection)                  
                    ws2[i].alignment = copy(cell1.alignment)            
            except AttributeError as e:             
                print("cell(%s) is %s" % (i,e))             
                continue    




excelFile = 'test.xlsx'
wb = openpyxl.load_workbook(excelFile)
sheet = wb.sheetnames
#excelFile2 = 'tess.xlsx'
#wb2 = openpyxl.load_workbook(excelFile2)

# 遍列Excel表中所有合并单元格
sh = wb[sheet[1]]
print(sh.merged_cells.ranges)
sh2 = wb[sheet[0]]


# [<CellRange A2:A5>, <CellRange A6:A9>, <CellRange A10:A11>]

for item in sh.merged_cells:
	print(item)
	#sh.unmerge_cells(str(item))	#取消单元格合并

#copy_entire_sheet_with_format(wb2,sh)
#copy_path_sheet_with_format(sh2, sh, 5,1,18,8)
copy_same_file_path_sheet_with_format(sh2, sh, 5,1,18,37,20,0)
wb.save('tess.xlsx')
